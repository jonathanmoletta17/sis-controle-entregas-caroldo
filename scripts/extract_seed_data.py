"""
Extract items per posto from CHECKLISTS CAROLDO.xlsx and generate a TypeScript seed file.
Output: /home/z/my-project/prisma/seed-data.ts
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import json
import re

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT_TS = "/home/z/my-project/prisma/seed-data.ts"

wb = openpyxl.load_workbook(SRC, data_only=True)

# Group: sheet -> category mapping
SHEETS = {
    'MATERIAIS MARCENEIROS':       ('Materiais', 'Marceneiro', 'Amarelo'),
    'MATERIAIS PEDREIROS':         ('Materiais', 'Pedreiro', 'Azul'),
    'MATERIAIS ELETRICISTAS':      ('Materiais', 'Eletricista', 'Laranja'),
    'MATERIAIS PINTORES':          ('Materiais', 'Pintor', 'Amarelo'),
    'MATERIAIS INST. HIDRÁULICOS': ('Materiais', 'Instalador Hidráulico', 'Azul'),
    'MATERIAIS TÉC. REFRIGERAÇÃO': ('Materiais', 'Técnico de Refrigeração', 'Amarelo'),
    'MATERIAIS TÉC. INFRA. DE REDES': ('Materiais', 'Técnico de Infraestrutura de Redes', 'Laranja'),
    'EPI SUPERVISOR':              ('EPI', 'Supervisor', None),
    'EPI MARCENEIRO':              ('EPI', 'Marceneiro', 'Amarelo'),
    'EPI ELETRICISTA':             ('EPI', 'Eletricista', 'Laranja'),
    'EPI SERVENTE':                ('EPI', 'Servente de Obras', 'Verde'),
    'EPI INST. HIDRÁULICO':        ('EPI', 'Instalador Hidráulico', 'Azul'),
    'EPI PEDREIRO':                ('EPI', 'Pedreiro', 'Azul'),
    'EPI PINTOR':                  ('EPI', 'Pintor', 'Amarelo'),
    'EPI TÉC. REFRIGERAÇÃO':       ('EPI', 'Técnico de Refrigeração', 'Amarelo'),
    'EPI TÉC. INFRA. DE REDES':    ('EPI', 'Técnico de Infraestrutura de Redes', 'Laranja'),
    'UNIFORME':                    ('Uniforme', None, None),  # uniform applies to all postos
    'UNIFORME FABIANO':            ('Uniforme', None, None),
    'UNIFORME ELETRICISTA':        ('Uniforme', None, None),
    'DOCUMENTOS':                  ('Documento', None, None),  # applies to all
}

def get_blocks(ws):
    blocks = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col > 1:
            blocks.append((mr.min_col, mr.max_col))
    has_a = any(b[0] == 1 for b in blocks)
    if not has_a:
        for mr in ws.merged_cells.ranges:
            if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col == 1:
                blocks.append((mr.min_col, mr.max_col))
                break
    blocks.sort()
    return blocks

def find_header_row(ws, sc, ec):
    for r in range(13, 25):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "Descrição" in v:
                return r
    return None

def get_unique_items(ws, sc, ec, hr):
    items = []
    last_desc = None
    for r in range(hr + 1, ws.max_row + 1):
        v = ws.cell(r, sc).value
        if v and isinstance(v, str):
            v_upper = v.upper()
            if "FISCALIZAÇÃO" in v_upper or "REPRESENTANTE" in v_upper or "ASSINATURAS" in v_upper or "DOCUMENTOS LISTADOS" in v_upper:
                break
        if v and isinstance(v, str) and v != last_desc and len(v.strip()) > 3:
            last_desc = v
            # Clean description
            desc = v.strip()
            # Skip if it looks like a date or number
            if re.match(r'^\d', desc) and len(desc) < 12:
                continue
            items.append(desc)
    return items

# Collect items per (category, posto)
items_by_posto = {}  # (categoria, posto) -> [items]
all_postos = set()
all_colaboradores_raw = []  # list of {nome, posto_hint, cpf_placeholder, data_admissao}

# For each sheet, get items from the FIRST block (template)
for sname, (categoria, posto, cor) in SHEETS.items():
    if sname not in wb.sheetnames:
        print(f"  WARN: sheet not found: {sname}")
        continue
    ws = wb[sname]
    blocks = get_blocks(ws)
    if not blocks:
        # UNIFORME FABIANO has no top-level blocks — use columns A-K
        blocks = [(1, 11)]
    sc, ec = blocks[0]
    hr = find_header_row(ws, sc, ec)
    if not hr:
        print(f"  WARN: header not found in {sname}")
        continue
    items = get_unique_items(ws, sc, ec, hr)
    if posto:
        key = (categoria, posto)
        if key not in items_by_posto:
            items_by_posto[key] = []
        # Deduplicate (some items appear multiple times)
        for it in items:
            if it not in items_by_posto[key]:
                items_by_posto[key].append(it)
        all_postos.add((posto, cor))
    else:
        # Uniforme and Documento apply to ALL postos
        for p, _ in list(all_postos):
            key = (categoria, p)
            if key not in items_by_posto:
                items_by_posto[key] = []
            for it in items:
                if it not in items_by_posto[key]:
                    items_by_posto[key].append(it)

# Build the TypeScript file
lines = []
lines.append("// Auto-generated seed data from CHECKLISTS CAROLDO.xlsx")
lines.append("// Source: /home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx")
lines.append("")
lines.append("export const POSTOS = [")
for nome, cor in sorted(all_postos):
    lines.append(f'  {{ nome: "{nome}", corCapacete: "{cor or ""}" }},')
lines.append("] as const;")
lines.append("")
lines.append("export const CATEGORIAS = [")
lines.append('  { nome: "Materiais", descricao: "Ferramentas e materiais permanentes por posto" },')
lines.append('  { nome: "EPI", descricao: "Equipamentos de Proteção Individual por posto" },')
lines.append('  { nome: "Uniforme", descricao: "Uniformes entregues aos terceirizados" },')
lines.append('  { nome: "Documento", descricao: "Documentação admissional e legal" },')
lines.append("] as const;")
lines.append("")
lines.append("// Itens por (categoria, posto) — lista de descrições")
lines.append("export const ITENS_POR_POSTO: Record<string, string[]> = {")
for (categoria, posto), items in sorted(items_by_posto.items()):
    key = f"{categoria}__{posto}"
    lines.append(f'  "{key}": [')
    for it in items:
        # Escape quotes and special chars
        safe = it.replace('\\', '\\\\').replace('"', '\\"').replace('\n', ' ').replace('\r', '').strip()
        if safe:
            lines.append(f'    "{safe}",')
    lines.append('  ],')
lines.append("};")
lines.append("")
# Add known collaborators extracted from the Excel (with normalized names and placeholder CPFs)
lines.append("// Colaboradores identificados no Excel (grafias normalizadas)")
lines.append("// CPFs são placeholders — devem ser preenchidos pelo gestor ao confirmar o cadastro")
lines.append("export const COLABORADORES_INICIAIS = [")
colabs = [
    ("Ivo Leandro da Silva Vieira", "Marceneiro", "2025-03-12", None),
    ("Andrea dos Santos Pavão", "Pedreiro", "2025-12-03", None),
    ("Joel Turela Tompsen", "Técnico de Infraestrutura de Redes", "2025-04-07", None),
    ("Gilberto Texeira da Rosa", "Pintor", "2025-10-13", None),
    ("Anderson Claiton Lopes da Cruz", "Eletricista", "2025-03-11", None),
    ("Dilmar Mendes de Souza", "Eletricista", "2025-10-06", None),
    ("Marcos Vinícius Machado da Silva", "Técnico de Refrigeração", "2025-03-12", None),
    ("Rogerio Alves", "Eletricista", "2025-12-03", None),
    ("José Fernando Neves Faller", "Eletricista", "2025-10-06", None),
    ("Fabiano L. Alves", "Pedreiro", "2025-12-03", None),
]
for i, (nome, posto, adm, deslig) in enumerate(colabs, 1):
    cpf_placeholder = f"000.000.000-{i:02d}"
    deslig_str = f'"{deslig}"' if deslig else "null"
    lines.append(f'  {{ nomeCompleto: "{nome}", cpf: "{cpf_placeholder}", posto: "{posto}", dataAdmissao: "{adm}", dataDesligamento: {deslig_str} }},')
lines.append("] as const;")
lines.append("")

Path(OUT_TS).write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {OUT_TS}")
print(f"  Postos: {len(all_postos)}")
print(f"  Categorias: 4")
print(f"  Combinações (categoria, posto): {len(items_by_posto)}")
print(f"  Total itens únicos (somando todas combinações): {sum(len(v) for v in items_by_posto.values())}")
print(f"  Colaboradores iniciais: {len(colabs)}")
