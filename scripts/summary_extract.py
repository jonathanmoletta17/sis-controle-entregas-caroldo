"""
Refined extraction:
- For each sheet, list UNIQUE items (deduplicate rows that are part of the same merged description cell)
- List all collaborator names found
- Count items per block
- Identify inconsistencies (blank names, blank dates, divergent contract numbers)
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
from collections import defaultdict

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT = "/home/z/my-project/scripts/_summary.txt"

wb = openpyxl.load_workbook(SRC, data_only=True)

lines = []
lines.append("=" * 100)
lines.append("SUMMARY: UNIQUE ITEMS + COLLABORATORS PER SHEET")
lines.append("=" * 100)

def get_blocks(ws):
    blocks = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col > 1:
            blocks.append((mr.min_col, mr.max_col))
    has_a = any(b[0] == 1 for b in blocks)
    if not has_a:
        for mr in ws.merged_cells.ranges:
            if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col == 1:
                blocks.append((mr.min_col, mr.max_col))
                break
    blocks.sort()
    return blocks

def find_header_row(ws, sc, ec):
    for r in range(13, 25):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "Descrição" in v:
                return r
    return None

def find_name(ws, sc, ec):
    """Return (name, name_row, name_col) or (None, None, None)."""
    for r in range(13, 20):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "NOME DO COLABORADOR" in v.upper():
                # Extract name after the colon
                name = re.sub(r'^.*NOME DO COLABORADOR\s*[:：]?\s*', '', v, flags=re.IGNORECASE).strip()
                name = re.sub(r'^[\.:\s]+|[\.:\s]+$', '', name)
                return name, r, c
            if v and isinstance(v, str) and r == 13 and "ESTADO" not in v and "Descrição" not in v and "CONTRATO" not in v and "CONTROLE" not in v and "SECRETARIA" not in v and "POSTO" not in v and "UNIDADE" not in v and "DIVISÃO" not in v and len(v) < 60:
                return v.strip(), r, c
    return None, None, None

def find_contract(ws, sc, ec):
    """Find contract number in the header block."""
    for r in range(1, 15):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "CONTRATO" in v.upper() and "MANUTENÇÃO" in v.upper():
                # Extract contract number like "003/2025" or "003/2026"
                m = re.search(r'(\d{3}/\d{4})', v)
                if m:
                    return m.group(1)
                return v
    return None

def find_footer_signatures(ws, sc, ec):
    """Find FISCALIZAÇÃO and REPRESENTANTE signatures."""
    sigs = {}
    for r in range(ws.max_row - 15, ws.max_row + 1):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str):
                if "FISCALIZAÇÃO" in v.upper():
                    # Look for the other signature in same row
                    for c2 in range(sc, ec + 1):
                        if c2 == c: continue
                        v2 = ws.cell(r, c2).value
                        if v2 and isinstance(v2, str) and "REPRESENTANTE" in v2.upper():
                            # Extract company name
                            m = re.search(r'REPRESENTANTE\s+(\w+)', v2.upper())
                            sigs['fiscalizacao'] = v[:80]
                            sigs['representante'] = v2[:80]
                            return sigs
                    sigs['fiscalizacao'] = v[:80]
    return sigs

def get_unique_items(ws, sc, ec, hr):
    """Get unique items by scanning merged description cells."""
    items = []
    # Find the description column = sc (first column of block)
    # Walk through rows, track when description cell changes (i.e., new item)
    last_desc = None
    for r in range(hr + 1, ws.max_row + 1):
        v = ws.cell(r, sc).value
        # Stop on footer
        if v and isinstance(v, str):
            v_upper = v.upper()
            if "FISCALIZAÇÃO" in v_upper or "REPRESENTANTE" in v_upper or "ASSINATURAS" in v_upper or "DOCUMENTOS LISTADOS" in v_upper:
                break
        # Find unit/date
        unit_v = None
        date_v = None
        # Determine unit and date columns by scanning header row
        # Use merged cell ranges to find them
        for c in range(sc, ec + 1):
            header_v = ws.cell(hr, c).value
            if not header_v:
                # Check if it's part of merged cell with header at top-left
                for mr in ws.merged_cells.ranges:
                    if mr.min_row <= hr <= mr.max_row and mr.min_col <= c <= mr.max_col:
                        header_v = ws.cell(mr.min_row, mr.min_col).value
                        break
            if header_v and isinstance(header_v, str):
                if "Uni" in header_v and not unit_v:
                    unit_v = ws.cell(r, c).value
                elif "Data" in header_v and not date_v:
                    date_v = ws.cell(r, c).value

        # If this row has a description value AND it's not the same as last (i.e., not part of merged cell continuing)
        if v and isinstance(v, str) and v != last_desc:
            last_desc = v
            date_str = ""
            if date_v:
                if hasattr(date_v, 'strftime'):
                    date_str = date_v.strftime("%d/%m/%Y")
                else:
                    date_str = str(date_v)
            items.append({
                "row": r,
                "desc": v,
                "unit": str(unit_v) if unit_v else "",
                "date": date_str,
            })
        elif v and isinstance(v, str) and v == last_desc:
            # Same description continuing in merged cell — skip (we already added it)
            pass
        # If no description but unit/date present, this is a "sub-row" within the same merged cell — skip for unique count
    return items

# Aggregate
all_summary = []

for sname in wb.sheetnames:
    ws = wb[sname]
    lines.append("\n" + "#" * 90)
    lines.append(f"SHEET: {sname}")
    lines.append(f"  dim={ws.dimensions}  rows={ws.max_row}  cols={ws.max_column}")
    lines.append("#" * 90)

    blocks = get_blocks(ws)
    lines.append(f"Blocks: {len(blocks)}")

    sheet_summary = {
        "sheet": sname,
        "blocks": [],
        "all_names": [],
    }

    for bi, (sc, ec) in enumerate(blocks, 1):
        # Find name
        name, name_r, name_c = find_name(ws, sc, ec)
        # Find contract
        contract = find_contract(ws, sc, ec)
        # Find header row
        hr = find_header_row(ws, sc, ec)
        # Find signatures
        sigs = find_footer_signatures(ws, sc, ec)
        # Get unique items
        items = get_unique_items(ws, sc, ec, hr) if hr else []

        # Count items with delivery date filled
        items_with_date = sum(1 for it in items if it["date"] and it["date"] not in ["-", "—"])

        block_info = {
            "block": bi,
            "cols": f"{get_column_letter(sc)}-{get_column_letter(ec)}",
            "name": name if name else "(vazio)",
            "contract": contract,
            "header_row": hr,
            "items_total": len(items),
            "items_with_date": items_with_date,
            "signatures": sigs,
        }
        sheet_summary["blocks"].append(block_info)
        if name:
            sheet_summary["all_names"].append(name)

        lines.append(f"\n  Block {bi} ({get_column_letter(sc)}-{get_column_letter(ec)}):")
        lines.append(f"    Colaborador: '{name if name else '(vazio)'}'")
        lines.append(f"    Contrato: {contract}")
        lines.append(f"    Header row: {hr}")
        lines.append(f"    Itens únicos: {len(items)}")
        lines.append(f"    Itens com data preenchida: {items_with_date}")
        if sigs:
            lines.append(f"    Assinaturas: {sigs}")
        # Show first 5 items
        if items:
            lines.append(f"    Primeiros 5 itens:")
            for i, it in enumerate(items[:5], 1):
                desc = it["desc"][:90] + "..." if len(it["desc"]) > 90 else it["desc"]
                lines.append(f"      {i}. unid={it['unit']} data={it['date']:12} | {desc}")
            if len(items) > 5:
                lines.append(f"      ... (+{len(items)-5} itens)")

    all_summary.append(sheet_summary)

# Final aggregated summary
lines.append("\n\n" + "=" * 90)
lines.append("AGGREGATED SUMMARY")
lines.append("=" * 90)
lines.append(f"\nTotal de abas: {len(all_summary)}")

# All unique names found across all sheets
all_names_set = set()
names_by_sheet = {}
for s in all_summary:
    sheet_names = []
    for b in s["blocks"]:
        if b["name"] and b["name"] != "(vazio)":
            all_names_set.add(b["name"])
            sheet_names.append(b["name"])
    if sheet_names:
        names_by_sheet[s["sheet"]] = sheet_names

lines.append(f"\nTotal de nomes únicos de colaboradores encontrados: {len(all_names_set)}")
lines.append("\nNomes por aba:")
for sheet, names in names_by_sheet.items():
    lines.append(f"  {sheet}: {names}")

lines.append("\nLista consolidada de terceirizados:")
for n in sorted(all_names_set):
    lines.append(f"  - {n}")

# Inconsistencies
lines.append("\n\n" + "=" * 90)
lines.append("INCONSISTÊNCIAS IDENTIFICADAS")
lines.append("=" * 90)
for s in all_summary:
    sheet_issues = []
    for b in s["blocks"]:
        if b["name"] == "(vazio)":
            sheet_issues.append(f"Block {b['block']} ({b['cols']}) sem nome de colaborador")
        if b["items_total"] > 0 and b["items_with_date"] == 0:
            sheet_issues.append(f"Block {b['block']} ({b['cols']}) tem {b['items_total']} itens mas NENHUMA data preenchida")
        if b["contract"] and "2026" in str(b["contract"]):
            sheet_issues.append(f"Block {b['block']} ({b['cols']}) usa contrato {b['contract']} (divergente do padrão 003/2025)")
    if sheet_issues:
        lines.append(f"\n{s['sheet']}:")
        for issue in sheet_issues:
            lines.append(f"  ⚠ {issue}")

Path(OUT).write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {len(lines)} lines to {OUT}")
print(f"File size: {Path(OUT).stat().st_size:,} bytes")
