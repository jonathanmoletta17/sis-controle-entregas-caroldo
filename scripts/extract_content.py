"""
Deep content extraction:
For each sheet, identify the "blocks" (groups of columns for each collaborator),
extract: collaborator name, list of items (with description, unit, delivery date, observation).
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT = "/home/z/my-project/scripts/_content_dump.txt"

wb = openpyxl.load_workbook(SRC, data_only=True)

lines = []
lines.append("=" * 100)
lines.append(f"WORKBOOK CONTENT EXTRACTION: {SRC}")
lines.append("=" * 100)

# Helper: detect block boundaries from merged cells in the header row.
# Each "block" is a set of columns belonging to one collaborator's checklist.
# Strategy: find row 1 (institutional header) merged ranges; each one = one block.
# Then within that block, the column header (Descrição / Uni. / Imagem / Data / Obs / Assinatura)
# is in row 15 (Materiais) or row 17/19 (EPI/Uniforme/Documentos).
def get_blocks(ws):
    """Return list of (start_col, end_col) tuples representing each collaborator block."""
    # Find merged ranges that span from row 1 to ~row 12 (the institutional header block)
    blocks = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col > 1:
            # Skip if this is contained within another block (some headers have nested merges)
            # We want top-level blocks
            blocks.append((mr.min_col, mr.max_col, mr.min_row, mr.max_row))
    # Also include column A block (often A1:K12) if not captured
    has_a_block = any(b[0] == 1 for b in blocks)
    if not has_a_block:
        for mr in ws.merged_cells.ranges:
            if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col == 1:
                blocks.append((mr.min_col, mr.max_col, mr.min_row, mr.max_row))
                break
    blocks.sort()
    return blocks

def find_header_row(ws, block_start_col, block_end_col):
    """Find the row that contains 'Descrição' header."""
    for r in range(13, 25):
        for c in range(block_start_col, block_end_col + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "Descrição" in v:
                return r
    return None

def find_name_row(ws, block_start_col, block_end_col):
    """Find the row that contains 'NOME DO COLABORADOR:'."""
    for r in range(13, 20):
        for c in range(block_start_col, block_end_col + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "NOME DO COLABORADOR" in v.upper():
                return r, c, v
            # Some sheets (Materiais) just have the name in row 13
            if v and isinstance(v, str) and r == 13 and len(v) < 60 and "ESTADO" not in v and "Descrição" not in v:
                return r, c, v
    return None, None, None

def find_subheader_cols(ws, header_row, block_start_col, block_end_col):
    """Map each column within block to its subheader (Descrição, Uni., Imagem, Data, Observação, Assinatura)."""
    col_map = {}
    # Some headers span multiple columns; check merged cells
    for c in range(block_start_col, block_end_col + 1):
        v = ws.cell(header_row, c).value
        if v and isinstance(v, str):
            col_map[c] = v.strip()
    # If only the first col has value, the rest are continuation; find via merged cells
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= header_row <= mr.max_row and mr.min_col >= block_start_col and mr.max_col <= block_end_col:
            top_val = ws.cell(mr.min_row, mr.min_col).value
            if top_val:
                for c in range(mr.min_col, mr.max_col + 1):
                    col_map[c] = str(top_val).strip()
    return col_map

for sname in wb.sheetnames:
    ws = wb[sname]
    lines.append("\n\n" + "#" * 100)
    lines.append(f"SHEET: {sname}")
    lines.append(f"  dimensions={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
    lines.append("#" * 100)

    blocks = get_blocks(ws)
    lines.append(f"\nDetected {len(blocks)} collaborator block(s):")
    for i, (sc, ec, sr, er) in enumerate(blocks, 1):
        lines.append(f"  Block {i}: cols {get_column_letter(sc)}-{get_column_letter(ec)} (rows {sr}-{er})")

    for bi, (sc, ec, sr, er) in enumerate(blocks, 1):
        lines.append(f"\n--- BLOCK {bi} (cols {get_column_letter(sc)}-{get_column_letter(ec)}) ---")
        # Find name
        name_r, name_c, name_v = find_name_row(ws, sc, ec)
        if name_v:
            # Clean
            name_clean = str(name_v).replace("NOME DO COLABORADOR:", "").replace("NOME DO COLABORADOR", "").strip()
            # Remove trailing colons/dots
            name_clean = re.sub(r'^[\.:\s]+|[\.:\s]+$', '', name_clean)
            lines.append(f"  Name (row {name_r}, col {get_column_letter(name_c)}): '{name_clean}'")
        else:
            lines.append(f"  Name: NOT FOUND")

        # Find header row
        hr = find_header_row(ws, sc, ec)
        if not hr:
            lines.append(f"  Header row NOT FOUND — skipping items")
            continue
        lines.append(f"  Header row: {hr}")
        col_map = find_subheader_cols(ws, hr, sc, ec)
        # Show column structure
        lines.append(f"  Column structure:")
        for c in sorted(col_map.keys()):
            lines.append(f"    col {get_column_letter(c)}: '{col_map[c]}'")

        # Extract items: scan rows after header_row
        # Each item is a row that has a description in the first column of the block
        desc_col = sc  # description is the first column of the block
        # But check: in Materiais sheets, description starts at sc and may span to sc+3
        # The actual description value is in column sc (top of merged cell)
        items = []
        for r in range(hr + 1, ws.max_row + 1):
            # Check if this row contains a description in the desc_col
            v = ws.cell(r, desc_col).value
            # Find unit, date, observation by matching columns
            unit_v = None
            date_v = None
            obs_v = None
            for c, lbl in col_map.items():
                if "Uni" in lbl and not unit_v:
                    unit_v = ws.cell(r, c).value
                elif "Data" in lbl and not date_v:
                    date_v = ws.cell(r, c).value
                elif ("Observ" in lbl or "Assinat" in lbl) and not obs_v:
                    obs_v = ws.cell(r, c).value
            # Stop if we hit footer rows
            if v and isinstance(v, str) and ("FISCALIZAÇÃO" in v.upper() or "REPRESENTANTE" in v.upper() or "assinaturas acima" in v.lower() or "documentos listados" in v.lower()):
                break
            # If row has any content, capture it
            if v or unit_v or date_v or obs_v:
                # Convert date to string
                date_str = ""
                if date_v:
                    if hasattr(date_v, 'strftime'):
                        date_str = date_v.strftime("%d/%m/%Y")
                    else:
                        date_str = str(date_v)
                items.append({
                    "row": r,
                    "desc": str(v) if v else "",
                    "unit": str(unit_v) if unit_v else "",
                    "date": date_str,
                    "obs": str(obs_v) if obs_v else ""
                })

        lines.append(f"\n  Items found: {len(items)}")
        # Print all items (limited to first 50 to avoid huge dump)
        for i, it in enumerate(items[:50], 1):
            desc = it["desc"][:80] + "..." if len(it["desc"]) > 80 else it["desc"]
            lines.append(f"    {i:3}. R{it['row']:3} | unid={it['unit']:3} | data={it['date']:12} | obs={it['obs'][:30]:30} | {desc}")
        if len(items) > 50:
            lines.append(f"    ... and {len(items)-50} more items")

        # Extract signature block info
        for r in range(ws.max_row - 10, ws.max_row + 1):
            for c in range(sc, ec + 1):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and ("FISCALIZAÇÃO" in v.upper() or "REPRESENTANTE" in v.upper() or "assinaturas acima" in v.lower() or "documentos listados" in v.lower()):
                    lines.append(f"  Footer (R{r} C{get_column_letter(c)}): {v[:150]}")

Path(OUT).write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {len(lines)} lines to {OUT}")
print(f"File size: {Path(OUT).stat().st_size:,} bytes")
