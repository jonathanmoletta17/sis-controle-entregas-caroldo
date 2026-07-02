"""
Deep inspection of CHECKLISTS CAROLDO.xlsx
For each sheet, extract:
  - The full header block (rows 1-15) merged cells
  - The column header row (the row that contains the item names)
  - The first 5 sample data rows
  - The last 5 data rows
  - All merged cell ranges
  - Column widths
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import json

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT = "/home/z/my-project/scripts/_inspection_dump.txt"

wb = openpyxl.load_workbook(SRC, data_only=True)

lines = []
lines.append("=" * 100)
lines.append(f"WORKBOOK: {SRC}")
lines.append(f"SHEETS ({len(wb.sheetnames)}): {wb.sheetnames}")
lines.append("=" * 100)

for sname in wb.sheetnames:
    ws = wb[sname]
    lines.append("\n\n" + "#" * 100)
    lines.append(f"SHEET: {sname}")
    lines.append(f"  dims={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
    lines.append("#" * 100)

    # Merged cells (just the count and first 10)
    merged = list(ws.merged_cells.ranges)
    lines.append(f"\nMerged cells: {len(merged)} ranges")
    for m in sorted(merged, key=lambda x: (x.min_row, x.min_col))[:25]:
        lines.append(f"  {m}")

    # Print first 15 rows, ALL columns, with cell value (truncated)
    lines.append("\n--- FIRST 20 ROWS (header block + first data rows) ---")
    max_r = min(20, ws.max_row)
    max_c = ws.max_column
    for r in range(1, max_r + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None:
                row_vals.append("")
            else:
                s = str(v).replace("\n", " | ").strip()
                if len(s) > 35:
                    s = s[:32] + "..."
                row_vals.append(s)
        # Find last non-empty cell to truncate trailing empties
        last = max_c
        while last > 0 and not row_vals[last-1]:
            last -= 1
        if last == 0:
            lines.append(f"  R{r:3}: (empty)")
        else:
            cells = " | ".join(row_vals[:last])
            lines.append(f"  R{r:3}: {cells}")

    # Find the actual data row count (rows after header that contain data)
    # Print a few sample middle data rows
    lines.append("\n--- SAMPLE MIDDLE DATA ROWS ---")
    if ws.max_row > 25:
        for r in [25, 30, 35, 40, 50]:
            if r <= ws.max_row:
                row_vals = []
                for c in range(1, max_c + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        row_vals.append("")
                    else:
                        s = str(v).replace("\n", " | ").strip()
                        if len(s) > 35:
                            s = s[:32] + "..."
                        row_vals.append(s)
                last = max_c
                while last > 0 and not row_vals[last-1]:
                    last -= 1
                if last == 0:
                    lines.append(f"  R{r:3}: (empty)")
                else:
                    cells = " | ".join(row_vals[:last])
                    lines.append(f"  R{r:3}: {cells}")

    # Print last 5 rows
    lines.append("\n--- LAST 5 ROWS ---")
    for r in range(max(2, ws.max_row - 4), ws.max_row + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v is None:
                row_vals.append("")
            else:
                s = str(v).replace("\n", " | ").strip()
                if len(s) > 35:
                    s = s[:32] + "..."
                row_vals.append(s)
        last = max_c
        while last > 0 and not row_vals[last-1]:
            last -= 1
        if last == 0:
            lines.append(f"  R{r:3}: (empty)")
        else:
            cells = " | ".join(row_vals[:last])
            lines.append(f"  R{r:3}: {cells}")

Path(OUT).write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {len(lines)} lines to {OUT}")
print(f"File size: {Path(OUT).stat().st_size:,} bytes")
