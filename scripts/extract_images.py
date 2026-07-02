"""
Extrai todas as imagens do CHECKLISTS CAROLDO.xlsx e mapeia para o item (descricao) correspondente.
Salva em /home/z/my-project/public/uploads/itens/ com nome {slug-da-descricao}-{hash}.png
Gera JSON /home/z/my-project/scripts/_imagem_map.json mapeando descricao -> [lista de arquivos]
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
import json, re, hashlib, io
from PIL import Image as PILImage

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT_DIR = Path("/home/z/my-project/public/uploads/itens")
OUT_DIR.mkdir(parents=True, exist_ok=True)
MAP_FILE = Path("/home/z/my-project/scripts/_imagem_map.json")

wb = openpyxl.load_workbook(SRC)

# Helper: get blocks (collaborator columns) per sheet
def get_blocks(ws):
    blocks = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col > 1:
            blocks.append((mr.min_col, mr.max_col))
    has_a = any(b[0] == 1 for b in blocks)
    if not has_a:
        for mr in ws.merged_cells.ranges:
            if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col == 1:
                blocks.append((mr.min_col, mr.max_col))
                break
    blocks.sort()
    return blocks

def find_header_row(ws, sc, ec):
    for r in range(13, 25):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "Descrição" in v:
                return r
    return None

def slugify(s, n=40):
    """Cria slug seguro da descricao para nome de arquivo."""
    s = str(s).lower()
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    s = re.sub(r'-+', '-', s)
    return s[:n].rstrip('-') or 'item'

# Para cada sheet, mapear imagens para o item (descricao) da mesma linha
descricao_to_images = {}  # {descricao_normalizada: [caminhos]}

for sname in wb.sheetnames:
    ws = wb[sname]
    if not hasattr(ws, '_images') or not ws._images:
        continue
    
    blocks = get_blocks(ws)
    if not blocks:
        # UNIFORME FABIANO tem layout diferente - tratar como único bloco A-K
        blocks = [(1, 11)]
    
    # Construir mapa: para cada bloco, ler descricao de cada item e registrar em qual linha está
    # Mapear (coluna, linha) -> descricao
    item_at_position = {}  # {(row, col): descricao}
    for sc, ec in blocks:
        hr = find_header_row(ws, sc, ec)
        if not hr:
            continue
        last_desc = None
        for r in range(hr + 1, ws.max_row + 1):
            v = ws.cell(r, sc).value
            if v and isinstance(v, str):
                v_upper = v.upper()
                if "FISCALIZAÇÃO" in v_upper or "REPRESENTANTE" in v_upper or "ASSINATURAS" in v_upper:
                    break
                if len(v.strip()) > 3 and v != last_desc:
                    last_desc = v.strip()
                    # Para cada coluna do bloco, marcar a posicao (r, c) -> descricao
                    for c in range(sc, ec + 1):
                        item_at_position[(r, c)] = last_desc

    # Para cada imagem, descobrir em qual linha/coluna ela está
    for img in ws._images:
        try:
            # anchor: top_left_cell (row, col) - 1-indexed
            anchor = img.anchor
            if hasattr(anchor, '_from') and anchor._from:
                row = anchor._from.row + 1  # 1-indexed
                col = anchor._from.col + 1  # 1-indexed
            elif hasattr(anchor, 'tl_cell'):
                # CellAnchor
                cell = anchor.tl_cell
                row = cell.row
                col = cell.col_idx if hasattr(cell, 'col_idx') else cell.column
            else:
                continue
            
            # Procurar item que contém essa posicao
            # A imagem geralmente está dentro das colunas "Imagem Ilustrativa" do bloco
            # Procurar pela posicao exata ou pelas proximidades (mesma linha)
            descricao = None
            for (r, c), d in item_at_position.items():
                if r == row:
                    # mesma linha - provavelmente é imagem deste item
                    descricao = d
                    break
                if r <= row <= r + 5:
                    descricao = d
                    # não break - preferir match exato de linha
            
            if not descricao:
                continue
            
            # Extrair a imagem
            img_data = img._data() if callable(getattr(img, '_data', None)) else img._data
            if not img_data:
                continue
            
            # Ler como bytes
            if hasattr(img_data, 'read'):
                img_bytes = img_data.read()
            elif hasattr(img_data, 'getvalue'):
                img_bytes = img_data.getvalue()
            elif isinstance(img_data, bytes):
                img_bytes = img_data
            else:
                continue
            
            if not img_bytes:
                continue
            
            # Hash para nome único
            hash8 = hashlib.md5(img_bytes).hexdigest()[:8]
            slug = slugify(descricao)
            
            # Detectar extensão
            try:
                pil = PILImage.open(io.BytesIO(img_bytes))
                ext = 'png' if pil.format == 'PNG' else 'jpg' if pil.format in ('JPEG', 'JPG') else 'png'
            except:
                ext = 'png'
            
            fname = f"{slug}-{hash8}.{ext}"
            fpath = OUT_DIR / fname
            
            # Evitar duplicar mesma imagem
            if not fpath.exists():
                fpath.write_bytes(img_bytes)
            
            # Mapear descricao -> arquivo
            if descricao not in descricao_to_images:
                descricao_to_images[descricao] = []
            if fname not in descricao_to_images[descricao]:
                descricao_to_images[descricao].append(f"/uploads/itens/{fname}")
        except Exception as e:
            continue

# Salvar mapa JSON
MAP_FILE.write_text(json.dumps(descricao_to_images, ensure_ascii=False, indent=2), encoding='utf-8')

print(f"✓ {len(descricao_to_images)} descricoes com imagens mapeadas")
print(f"✓ {len(list(OUT_DIR.glob('*')))} arquivos salvos em {OUT_DIR}")
print(f"✓ Mapa salvo em {MAP_FILE}")

# Amostra
for i, (desc, imgs) in enumerate(list(descricao_to_images.items())[:5]):
    print(f"  {i+1}. {desc[:60]}... → {len(imgs)} imagem(s)")
