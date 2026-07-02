"""
Refinado: cada imagem corresponde a 1 (e apenas 1) item.
Estratégia:
- Para cada bloco lateral, conhecer a estrutura: descrição ocupa as 4 primeiras colunas (mescladas verticalmente),
  Imagem Ilustrativa ocupa as 3 colunas seguintes (mescladas verticalmente), Data de entrega ocupa 3 colunas, etc.
- Quando uma imagem está em colunas "Imagem Ilustrativa" (5ª-7ª colunas do bloco), associar à descrição do item
  na mesma linha (ou na linha da célula mesclada superior, se estiver dentro de um range mesclado verticalmente).
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import json, re, hashlib, io
from PIL import Image as PILImage

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT_DIR = Path("/home/z/my-project/public/uploads/itens")
OUT_DIR.mkdir(parents=True, exist_ok=True)
MAP_FILE = Path("/home/z/my-project/scripts/_imagem_map.json")

wb = openpyxl.load_workbook(SRC)

def get_blocks(ws):
    blocks = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col > 1:
            blocks.append((mr.min_col, mr.max_col))
    has_a = any(b[0] == 1 for b in blocks)
    if not has_a:
        for mr in ws.merged_cells.ranges:
            if mr.min_row == 1 and mr.max_row >= 10 and mr.min_col == 1:
                blocks.append((mr.min_col, mr.max_col))
                break
    blocks.sort()
    return blocks

def find_header_row(ws, sc, ec):
    for r in range(13, 25):
        for c in range(sc, ec + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "Descrição" in v:
                return r
    return None

def slugify(s, n=40):
    s = str(s).lower()
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    s = re.sub(r'-+', '-', s)
    return s[:n].rstrip('-') or 'item'

def find_merged_top(ws, row, col):
    """Se a célula (row, col) faz parte de um merged range, retorna (top_row, top_col, bottom_row)."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col, mr.max_row
    return row, col, row

# Para cada bloco de cada aba, descobrir:
# 1. As colunas de "Imagem Ilustrativa" (sc+4 até sc+6)
# 2. As colunas de "Descrição" (sc até sc+3)
# 3. Para cada imagem, achar a célula superior do merged range vertical onde ela caiu
#    e achar o merged range equivalente na coluna de descrição (mesmo range de linhas)

descricao_to_image = {}  # {descricao: caminho}

for sname in wb.sheetnames:
    ws = wb[sname]
    if not hasattr(ws, '_images') or not ws._images:
        continue
    
    blocks = get_blocks(ws)
    if not blocks:
        blocks = [(1, 11)]
    
    for sc, ec in blocks:
        hr = find_header_row(ws, sc, ec)
        if not hr:
            continue
        
        # Identificar colunas de descrição e imagem pelo header row
        desc_cols = []  # colunas onde o header é "Descrição..."
        img_cols = []
        for c in range(sc, ec + 1):
            v = ws.cell(hr, c).value
            if v and isinstance(v, str):
                if "Descrição" in v:
                    desc_cols.append(c)
                elif "Imagem" in v or "IMAGEM" in v:
                    img_cols.append(c)
        
        if not desc_cols or not img_cols:
            continue
        
        # Construir mapa: para cada merged range que cobre uma coluna de imagem, qual é a descricao?
        # Estratégia: para cada imagem, descobrir o top_row da célula onde ela está (na coluna de imagem)
        # e achar o merged range na coluna de descrição que cobre esse top_row
        for img in ws._images:
            try:
                anchor = img.anchor
                if hasattr(anchor, '_from') and anchor._from:
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
                else:
                    continue
                
                # Verificar se a coluna é uma das colunas de imagem deste bloco
                if col not in img_cols:
                    continue
                
                # Achar top_row da célula da imagem (se for merged, pega o topo)
                img_top_row, _, img_bottom_row = find_merged_top(ws, row, col)
                
                # Achar merged range na coluna de descrição que cobre img_top_row
                desc_col = desc_cols[0]  # primeira coluna de descrição
                desc_top_row, _, desc_bottom_row = find_merged_top(ws, img_top_row, desc_col)
                
                # A descrição está na célula (desc_top_row, desc_col)
                desc_val = ws.cell(desc_top_row, desc_col).value
                if not desc_val or not isinstance(desc_val, str):
                    continue
                desc_val = desc_val.strip()
                if len(desc_val) < 3:
                    continue
                if "FISCALIZAÇÃO" in desc_val.upper() or "REPRESENTANTE" in desc_val.upper():
                    continue
                
                # Extrair a imagem
                img_data = img._data() if callable(getattr(img, '_data', None)) else img._data
                if not img_data:
                    continue
                if hasattr(img_data, 'read'):
                    img_bytes = img_data.read()
                elif hasattr(img_data, 'getvalue'):
                    img_bytes = img_data.getvalue()
                elif isinstance(img_data, bytes):
                    img_bytes = img_data
                else:
                    continue
                if not img_bytes:
                    continue
                
                # Hash + slug para nome único
                hash8 = hashlib.md5(img_bytes).hexdigest()[:8]
                slug = slugify(desc_val)
                
                try:
                    pil = PILImage.open(io.BytesIO(img_bytes))
                    ext = 'png' if pil.format == 'PNG' else 'jpg' if pil.format in ('JPEG', 'JPG') else 'png'
                except:
                    ext = 'png'
                
                fname = f"{slug}-{hash8}.{ext}"
                fpath = OUT_DIR / fname
                
                if not fpath.exists():
                    fpath.write_bytes(img_bytes)
                
                # Mapear descricao -> arquivo (apenas 1 imagem por item; se já existe, manter a primeira)
                if desc_val not in descricao_to_image:
                    descricao_to_image[desc_val] = f"/uploads/itens/{fname}"
            except Exception:
                continue

# Salvar mapa
MAP_FILE.write_text(json.dumps(descricao_to_image, ensure_ascii=False, indent=2), encoding='utf-8')

print(f"✓ {len(descricao_to_image)} itens com imagem (1 imagem cada)")
arqs = list(OUT_DIR.glob('*'))
print(f"✓ {len(arqs)} arquivos em {OUT_DIR}")
print(f"✓ Mapa em {MAP_FILE}")
print()
print("Amostra:")
for i, (desc, img) in enumerate(list(descricao_to_image.items())[:5]):
    print(f"  {i+1}. {desc[:60]}...")
    print(f"     → {img}")
