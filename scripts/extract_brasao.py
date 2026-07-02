"""Extrai o brasão do RS do Excel (provavelmente embutido como imagem no header institucional)."""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import hashlib, io
from PIL import Image

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
OUT_DIR = Path("/home/z/my-project/public")
wb = openpyxl.load_workbook(SRC)

# Procurar imagens em cada sheet — o brasão provavelmente está no header (linhas 1-12)
# de uma das primeiras abas (geralmente na primeira aba de materiais)
brasoes_candidatos = []

for sname in wb.sheetnames:
    ws = wb[sname]
    if not hasattr(ws, '_images') or not ws._images:
        continue
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from') and anchor._from:
                row = anchor._from.row + 1
                col = anchor._from.col + 1
            else:
                continue
            # Imagem no header institucional (linhas 1-12) é candidata a brasão
            if row <= 12:
                img_data = img._data() if callable(getattr(img, '_data', None)) else img._data
                if not img_data:
                    continue
                if hasattr(img_data, 'read'):
                    img_bytes = img_data.read()
                elif hasattr(img_data, 'getvalue'):
                    img_bytes = img_data.getvalue()
                elif isinstance(img_data, bytes):
                    img_bytes = img_data
                else:
                    continue
                if not img_bytes:
                    continue
                # Tamanho típico de brasão: entre 5KB e 500KB
                size = len(img_bytes)
                if size < 1000:
                    continue
                # Detectar formato
                try:
                    pil = Image.open(io.BytesIO(img_bytes))
                    w, h = pil.size
                    fmt = pil.format
                    # Brasão geralmente é quadrado ou quase (ratio 0.7-1.3)
                    ratio = w / h if h > 0 else 0
                    if 0.5 < ratio < 1.5 and w >= 50 and h >= 50:
                        brasoes_candidatos.append({
                            'sheet': sname,
                            'row': row,
                            'col': col,
                            'size_bytes': size,
                            'width': w,
                            'height': h,
                            'format': fmt,
                            'img_bytes': img_bytes,
                            'hash': hashlib.md5(img_bytes).hexdigest()[:8],
                        })
                except Exception:
                    continue
        except Exception:
            continue

print(f"Candidatos a brasão encontrados: {len(brasoes_candidatos)}")
# Mostrar candidatos únicos por hash
unicos = {}
for b in brasoes_candidatos:
    if b['hash'] not in unicos:
        unicos[b['hash']] = b
print(f"Únicos por hash: {len(unicos)}")
for h, b in unicos.items():
    print(f"  hash={h} sheet={b['sheet']} row={b['row']} col={b['col']} {b['width']}x{b['height']} {b['format']} {b['size_bytes']}B")

# Salvar todos os candidatos únicos para inspeção visual
OUT_DIR.mkdir(parents=True, exist_ok=True)
for i, (h, b) in enumerate(unicos.items(), 1):
    ext = 'png' if b['format'] == 'PNG' else 'jpg' if b['format'] in ('JPEG', 'JPG') else 'png'
    fname = f"_brasao_candidato_{i}_{h}.{ext}"
    (OUT_DIR / fname).write_bytes(b['img_bytes'])
    print(f"  Salvo: {fname}")
