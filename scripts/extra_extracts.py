"""Get UNIFORME FABIANO and MATERIAIS PEDREIROS items + other missing details."""
import openpyxl
from openpyxl.utils import get_column_letter
import re

SRC = "/home/z/my-project/upload/CHECKLISTS CAROLDO.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

# UNIFORME FABIANO
ws = wb["UNIFORME FABIANO"]
print("=== UNIFORME FABIANO ===")
print(f"dim={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
# Header at A2:K13, name at A16
for r in [2, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]:
    row_vals = []
    for c in range(1, 12):
        v = ws.cell(r, c).value
        if v is not None:
            s = str(v).replace("\n", " | ")
            if len(s) > 60: s = s[:57] + "..."
            row_vals.append(f"{get_column_letter(c)}={s}")
    if row_vals:
        print(f"  R{r}: {' | '.join(row_vals)}")

# Walk all items starting at row 21 (header is at row 20)
print("\nItens (a partir da linha 21):")
items = []
for r in range(21, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v and isinstance(v, str) and ("FISCALIZAÇÃO" in v.upper() or "REPRESENTANTE" in v.upper() or "assinaturas acima" in v.lower()):
        print(f"  R{r} (FOOTER): {v[:80]}")
        break
    if v and isinstance(v, str):
        items.append((r, v[:80]))
print(f"Total itens: {len(items)}")
for r, v in items[:15]:
    print(f"  R{r}: {v}")

# MATERIAIS PEDREIROS items
print("\n\n=== MATERIAIS PEDREIROS ===")
ws = wb["MATERIAIS PEDREIROS"]
# Items are in columns P-AC, starting row 17
items = []
for r in range(17, ws.max_row + 1):
    v = ws.cell(r, 16).value  # column P = 16
    if v and isinstance(v, str) and ("FISCALIZAÇÃO" in v.upper() or "REPRESENTANTE" in v.upper() or "assinaturas acima" in v.lower()):
        break
    if v and isinstance(v, str):
        items.append((r, v[:80]))
print(f"Total itens: {len(items)}")
for r, v in items:
    print(f"  R{r}: {v}")

# Also check the "zz" found in row 12 of MATERIAIS PEDREIROS
print("\nMATERIAIS PEDREIROS R12:")
for c in range(1, 35):
    v = ws.cell(12, c).value
    if v:
        print(f"  C{get_column_letter(c)}: {v}")

# Check the ELÉTRICA text in UNIFORME ELETRICISTA R13
print("\n\n=== UNIFORME ELETRICISTA R13 ===")
ws = wb["UNIFORME ELETRICISTA"]
for c in range(1, 26):
    v = ws.cell(13, c).value
    if v:
        print(f"  C{get_column_letter(c)}: {v}")

# Check for any extra names in all sheets — extract from row 13 of Materiais sheets
print("\n\n=== Names in row 13 of Materiais sheets ===")
for sname in ["MATERIAIS MARCENEIROS", "MATERIAIS PEDREIROS", "MATERIAIS ELETRICISTAS",
              "MATERIAIS PINTORES", "MATERIAIS INST. HIDRÁULICOS", "MATERIAIS TÉC. REFRIGERAÇÃO",
              "MATERIAIS TÉC. INFRA. DE REDES"]:
    ws = wb[sname]
    for r in [12, 13, 14]:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "ESTADO" not in v and "CONTRATO" not in v and "Descrição" not in v and "SECRETARIA" not in v and "UNIDADE" not in v and "DIVISÃO" not in v and "CONTROLE" not in v and "POSTO" not in v and "MARCENEIROS" not in v and "PEDREIROS" not in v and "ELETRICISTAS" not in v and "PINTORES" not in v and "INSTALADORES" not in v and "REFRIGERAÇÃO" not in v and "INFRAESTRUTURA" not in v and "TÉCNICOS" not in v and "TÉCNICO" not in v and len(v) < 80:
                print(f"  {sname} R{r} C{get_column_letter(c)}: {v}")
